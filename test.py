import os
import json
import torch
import torch.nn as nn
import numpy as np
import os.path as osp
import torch.nn.functional as F
from PIL import Image
from dataloader import CreateTestDataLoader
from models import CreateModel
from torch.autograd import Variable
from options.test_options import TestOptions
from utils.metric import Evaluator
from utils import *
import cv2
import matplotlib.pyplot as plt
import openpyxl

torch.set_num_threads(2)
torch.set_printoptions(precision=3)
np.set_printoptions(precision=3)


def test(num_classes, loader, model, ignore_index):
    real_perc = torch.zeros(num_classes)
    pred_perc = torch.zeros(num_classes)

    model.eval()
    metric = Evaluator(num_class=num_classes)
    with torch.no_grad():
        for j, (img, lbl, _) in enumerate(loader):
            img = img.cuda()      # to gpu
            b, _, h, w = img.size()

            pred = model(img).permute(0,2,3,1).reshape(-1, num_classes).cpu()  # N x K
            prob = F.softmax(pred, dim=-1)                                     # N x K
            lbl  = lbl.reshape(-1)                                             # N 
            max_prob, pl = torch.max(prob, dim=-1)           # N
            # mask_low = max_prob < thr

            metric.add_batch(lbl.cpu().detach().numpy(), pl.cpu().detach().numpy())

            # statistic
            for i in range(num_classes):
                real_perc[i] += (lbl==i).sum()
                pred_perc[i] += (pl==i).sum()

    acc = metric.Pixel_Accuracy()
    Precision, Recall = metric.Precision_Recall()
    IoU = np.around(metric.Intersection_over_Union()*100, 2)
    F1 = np.around(metric.F1_Score()*100, 2)
    Kappa = np.around(metric.Kappa(), 4)

    IoU = np.delete(IoU, ignore_index, axis=0)
    F1 = np.delete(F1, ignore_index, axis=0)

    real_perc = real_perc / real_perc.sum()
    pred_perc = pred_perc / pred_perc.sum()
    real_perc = torch.round(real_perc * 10**3) / 10**3
    pred_perc = torch.round(pred_perc * 10**3) / 10**3

    print ('real_perc: ', real_perc)
    print ('pred_perc: ', pred_perc)
    print ('IoU:       ', IoU, np.mean(IoU))
    print ('F1:        ', F1, np.mean(F1))
    print ('Kappa:     ', Kappa, np.mean(Kappa))
    print ('Precision: ', Precision)
    print ('Recall:    ', Recall)

    # for i in range(num_classes):
    #     mean = np.mean(pred_dist_gt[i].cpu().numpy(), axis=0)
    #     cov = np.cov(pred_dist_gt[i].cpu().numpy(), rowvar=False)

    #     print (i)
    #     print (mean)
    #     print ()

    return IoU, F1, Kappa


if __name__ == '__main__':
    opt = TestOptions()
    args = opt.initialize()
    data_dir = os.path.join(args.data_dir, args.dataset.replace('ISPRS_', 'ISPRS/'))
    # opt.print_options(args)
    os.environ['CUDA_VISIBLE_DEVICES'] = str(args.gpu)

    model = CreateModel(args.model, args.num_classes).cuda()

    # resume checkpoint
    save_dir = os.path.join(args.save_dir, args.model, args.method)
    checkpoint_file = '{}_{}%.pth'.format(args.dataset, args.percent)
    checkpoint_file = os.path.join(save_dir, checkpoint_file)

    resume = torch.load(checkpoint_file)
    model.load_state_dict(resume['state_dict'], strict=False)
    # epoch = resume['epoch']
    # best_acc = resume['best_acc']
    # best_mIoU = resume['best_mIoU']
    # print ('loaded checkpoint from: {}, epoch: {}, best_mIoU: {} '\
    #         .format(checkpoint_file, epoch, best_mIoU))

    print ('Tested...')
    test_loader = CreateTestDataLoader(data_dir, 'lists/test.txt', batch_size=args.batch_size)
    test_IoU, test_F1, test_Kappa = test(args.num_classes, test_loader, model, args.ignore_index)
    test_IoU = np.append(test_IoU, np.around(np.mean(test_IoU), 2))
    test_F1 = np.append(test_F1, np.around(np.mean(test_F1), 2))

    test_metric = []
    for idx in range(len(test_IoU)):
        if idx + 1 < len(test_IoU):
            test_metric.append(f"{test_IoU[idx]}/{test_F1[idx]}")
        else:
            test_metric.append(f"{test_IoU[idx]}/{test_F1[idx]}/{str(test_Kappa)}")
    test_metric = [args.dataset, str(args.percent), args.method] + test_metric

    excel_file = './outputs/test_metric.xlsx'
    if not os.path.isfile(excel_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.save(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    for col, data in enumerate(test_metric, 1):
        cell = sheet.cell(row=next_row, column=col)
        cell.value = data
    workbook.save(excel_file)

    print ()
    print ()
    print ()





